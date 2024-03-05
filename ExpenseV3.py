#version 3.2

import openpyxl
import re
from datetime import datetime

open_data = open("Bill.txt")
text_data = open_data.read()

# Define multiple extraction patterns
patterns = [
    r"Money Transfer:Rs ([0-9.]+) .*? on (\d{2}-\d{2}-\d{2})",
    r"HDFC Bank: Rs ([0-9.]+) .*? on (\d{2}-\d{2}-\d{2})",
    r"INR ([0-9,]+) spent on .*? on (\d{2}-[A-Za-z]{3}-\d{2})"
]

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Expenses"

# Set custom headers to the worksheet
headers = ["Date", "Amount"]
worksheet.append(headers)

# Add data to the worksheet for each pattern
data_rows = []
for pattern in patterns:
    matches = re.findall(pattern, text_data)
    for match in matches:
        amount, raw_date = match
        # Parse the raw date using datetime to reformat it to "dd/mm/yy" format
        date_obj = datetime.strptime(raw_date, "%d-%m-%y")
        data_rows.append([date_obj, float(amount)])  # Convert amount to float for proper sorting

# Sort the data by column A (Date) in ascending order
data_rows.sort(key=lambda x: x[0])

# Append the sorted data to the worksheet
for row in data_rows:
    formatted_date = row[0].strftime("%d/%m/%y")
    worksheet.append([formatted_date, row[1]])

# Calculate and add the sum of amounts to column B
total_amount_formula = f"=SUM(B2:B{worksheet.max_row})"
worksheet.cell(row=worksheet.max_row + 2, column=2, value=total_amount_formula).font = openpyxl.styles.Font(bold=True)
worksheet.cell(row=worksheet.max_row, column=1, value="Total").font = openpyxl.styles.Font(bold=True)

# Save the workbook to an XLSX file
workbook.save("Expenses.xlsx")

print("Data has been successfully extracted, sorted, and saved to Expenses.xlsx.")
